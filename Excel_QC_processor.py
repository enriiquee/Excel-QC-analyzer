# Import the necessary libraries
import subprocess

# Check if pandas is installed
try:
    import pandas as pd
except ImportError:
    # Install pandas using pip if not found
    subprocess.check_call(["python", "-m", "pip", "install", "pandas"])
    import pandas as pd

# Check if openpyxl is installed
try:
    import openpyxl
    # Import required modules from openpyxl
    from openpyxl.styles import Color, PatternFill, Font, Border
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule

except ImportError:
    # Install openpyxl using pip if not found
    subprocess.check_call(["python", "-m", "pip", "install", "openpyxl"])
    import openpyxl
    # Import required modules from openpyxl
    from openpyxl.styles import Color, PatternFill, Font, Border
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule

# Check if tkinter is installed
try:
    import tkinter as tk
    from tkinter import filedialog
    from tkinter.ttk import Progressbar
except ImportError:
    # Install tkinter using pip if not found
    subprocess.check_call(["python", "-m", "pip", "install", "tkinter"])
    import tkinter as tk
    from tkinter import filedialog
    from tkinter.ttk import Progressbar

# Check if threading is installed
try:
    import threading
except ImportError:
    # Install threading using pip if not found
    subprocess.check_call(["python", "-m", "pip", "install", "threading"])
    import threading


# Custom exception to be raised if the format of the Excel file is invalid
class InvalidFormatException(Exception):
    pass

# Function to check if the format of the Excel file is correct
def is_format_correct(df):
    columns_to_check = ["Scan name", "ROI name", "Segment name", "Tags", "QC status",
                        "Binding Density", "FoV registration QC", "Positive norm factor",
                        "Surface area", "Nuclei count", "QC flags"]
    for column in columns_to_check:
        if column not in df.columns:
            return False
    return True

def process_excel():
    thread = threading.Thread(target=process_excel_thread)
    thread.start()

# Function to process the Excel file in a separate thread
def process_excel_thread():
    global file_path
    if not file_path:
        output_text.set("Por favor, seleccione un archivo antes de hacer clic en Run.")
        return

    try:
        # Step 1: Read the Excel file
        update_progress(10)
        df = pd.read_excel(file_path)

        # Check if the Excel file has the correct format
        if not is_format_correct(df):
            raise InvalidFormatException("El archivo no contiene el formato esperado.")

        # Procesamiento de Excel
        columns_to_keep = ["Scan name", "ROI name", "Segment name", "Tags", "QC status",
                           "Binding Density", "FoV registration QC", "Positive norm factor",
                           "Surface area", "Nuclei count", "QC flags"]
        df_qc = df[columns_to_keep]

        df_qc_warning = df_qc[df_qc["QC status"].str.contains("WARNING")]

        with pd.ExcelWriter("QC_filter_test.xlsx", engine="openpyxl") as writer:
            df_qc.to_excel(writer, sheet_name="QC", index=False)
            df_qc_warning.to_excel(writer, sheet_name="QC_Warning", index=False)

            workbook = writer.book
            sheet = writer.sheets["QC_Warning"]
            sheet.insert_cols(sheet.max_column+1)
            sheet.cell(row=1, column=sheet.max_column+1, value="Considered as good one")

            qc_sheet = writer.sheets["QC"]
            qc_warning_sheet = writer.sheets["QC_Warning"]

            colors = {"PASS": "C6EFCE", "WARNING": "FFA07A"}

            qc_column = df_qc["QC status"]
            for i, qc_status in enumerate(qc_column):
                cell = f"E{i+2}"
                if qc_status in colors:
                    fill = PatternFill(start_color=colors[qc_status], end_color=colors[qc_status], fill_type="solid")
                    qc_sheet[cell].fill = fill

            qc_warning_column = df_qc_warning["QC status"]
            for i, qc_status in enumerate(qc_warning_column):
                cell = f"E{i+2}"
                if qc_status in colors:
                    fill = PatternFill(start_color=colors[qc_status], end_color=colors[qc_status], fill_type="solid")
                    qc_warning_sheet[cell].fill = fill

            for sheet in workbook.worksheets:
                sheet.sheet_format.defaultRowHeight = 15

            for sheet in workbook.worksheets:
                for i, column in enumerate(df_qc):
                    column_width = max(df_qc[column].astype(str).map(len).max(), len(column))
                    sheet.column_dimensions[get_column_letter(i+1)].width = column_width
                for i, row in enumerate(df_qc.index):
                    row_height = 15
                    sheet.row_dimensions[i + 2].height = row_height

            

        output_text.set("Archivo procesado con éxito.")
        update_progress(100)
        root.after(5000, close_app_after_delay)

    except InvalidFormatException as e:
        output_text.set(f"Error: {e}")
    except Exception as e:
        output_text.set(f"Error inesperado: {e}")
# Function to open the file dialog and select an Excel file
def open_file_dialog():
    global file_path
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    file_path_text.set(file_path)

# Function to update the progress bar
def update_progress(percentage):
    progress_bar["value"] = percentage
    root.update_idletasks()

# Function to close the app after a delay of 5 seconds
def close_app_after_delay():
    global root
    root.destroy()

# Function to run the app
def run_app():
    global root, file_path, file_path_text, output_text, progress_bar

    file_path = ""

    root = tk.Tk()
    root.geometry("800x400")
    root.title("Procesador de archivos de Excel")

    file_path_text = tk.StringVar()
    output_text = tk.StringVar()

    file_label = tk.Label(root, text="Archivo seleccionado:")
    file_label.pack(pady=10)

    file_entry = tk.Entry(root, textvariable=file_path_text, width=80)
    file_entry.pack()

    browse_button = tk.Button(root, text="Examinar", command=open_file_dialog)
    browse_button.pack(pady=10)

    run_button = tk.Button(root, text="Run", command=process_excel)
    run_button.pack(pady=10)

    output_label = tk.Label(root, textvariable=output_text)
    output_label.pack(pady=10)

    progress_bar = Progressbar(root, orient="horizontal", length=500, mode="determinate")
    progress_bar.pack(pady=10)

    root.mainloop()

if __name__ == "__main__":
    run_app()

